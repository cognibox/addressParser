import openpyxl
import argparse
import csv
# noinspection PyUnresolvedReferences,PyPackageRequirements
from postal.parser import parse_address

# define commandline parser
parser = argparse.ArgumentParser(
    description='Tool to extract addresses from blob in excel,'
                'all input/output files must be in the current directory',
    formatter_class=argparse.RawTextHelpFormatter)
parser.add_argument('address_list',
                    help=f'Excel file xlsx with addresses')

parser.add_argument('output',
                    help=f'csv file with the address decomposed in:\n'
                         f'street, city, state, postal code, country and analysis')

parser.add_argument('--column', dest='column', action='store',
                    default=1,
                    help='column index where the addresses are stored (default is 1)')

parser.add_argument('--output_encoding', dest='output_encoding', action='store',
                    default='utf-8-sig',
                    help='Encoding for the hc list (default: utf-8-sig)')

parser.add_argument('--no_headers', dest='no_headers', action='store_true',
                    help='to indicate that input file have no headers')

args = parser.parse_args()

data_path = './data/'
address_file = data_path + args.address_list
output_file = data_path + args.output

print(f'Reading address data file...')
wb_obj = openpyxl.load_workbook(address_file)
print(f'Reading address data file... DONE')
sheet_obj = wb_obj.active
max_row = sheet_obj.max_row
nbr_address = sheet_obj.max_row if args.no_headers else sheet_obj.max_row - 1
print(f'Number of adresses: {nbr_address}')
print(f'reading adresses from column: {args.column}')
with open(output_file, 'w', newline='', encoding=args.output_encoding) as result_file:
    writer = csv.writer(result_file)
    writer.writerow(['address', 'city', 'state', 'postal code', 'country', 'analysis'])
    start = 1 if args.no_headers else 2
    for i in range(start, sheet_obj.max_row+1):
        row = []
        address = {}
        street = []
        value = sheet_obj.cell(row=i, column=(int(args.column))).value
        address_parts = parse_address(value)
        for value, part in address_parts:
            if part in ('house_number', 'road', 'unit', 'po_box', 'city', 'state', 'postcode', 'country'):
                address[part] = value
        if 'house_number' in address:
            street.append(address['house_number'])
        if 'road' in address:
            if 'unit' in address:
                street.append(', '.join([address['road'], address['unit']]))
            else:
                street.append(address['road'])
        if 'po_box' in address:
            street.append(address['po_box'])
        street = " ".join(street) if street else ''
        writer.writerow([street, address.get("city", ""), address.get("state", ""),
                         address.get("postcode", ""), address.get("country", ""), str(address_parts)])
        print(f'{i-start+1} of {nbr_address}')
    print(f'Completed address extraction...')
